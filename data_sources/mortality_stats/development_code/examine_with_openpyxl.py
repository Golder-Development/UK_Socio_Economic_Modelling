from openpyxl import load_workbook
from pathlib import Path

ONS_DOWNLOADS = Path("ons_downloads/extracted")

# Examine ICD1 file
print("=" * 80)
print("EXAMINING ICD1.XLS")
print("=" * 80)

try:
    wb = load_workbook(ONS_DOWNLOADS / "icd1.xls", read_only=True, data_only=True)
    print(f"Sheet names: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames[:2]:  # First 2 sheets
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        
        # Print first row (headers)
        first_row = [cell.value for cell in ws[1]]
        print(f"Headers: {first_row}")
        
        # Print a few data rows
        print("\nFirst 5 data rows:")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
            print(f"Row {i}: {row}")
        
        print()
except Exception as e:
    print(f"Error loading with openpyxl: {e}")
    print("The file may be in old .xls format, need xlrd")
